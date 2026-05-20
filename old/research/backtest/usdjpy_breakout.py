# -*- coding: utf-8 -*-
"""
USD/JPY 1分足バックテスト — 20SMA接線ブレイクアウト戦略 v3

【戦略ロジック】
  準備: 20SMA接線が一定本数連続で平行（傾き < flat_threshold）
  エントリー:
    Long  = 準備OK × 当日始値より価格が上 × 価格がSMAを上抜け
    Short = 準備OK × 当日始値より価格が下 × 価格がSMAを下抜け
  損切: 直近スウィング安値（Long）/ 高値（Short）
  利確: 固定 10pips（0.10円）
  フィルター: イベント前後60分エントリー禁止

実行: python backtest/usdjpy_breakout.py
"""

from __future__ import annotations

import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import re
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional

import pandas as pd
import numpy as np

BASE_DIR   = Path(__file__).parent
PRICE_CSV  = BASE_DIR / "USDJPY" / "USDJPY__20241014~20260501（1分足）.csv"
EVENTS_XLS = BASE_DIR.parent / "米国経済指標カレンダー_2023Dec-2026May.xlsx"
OUTPUT_DIR = BASE_DIR / "results"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

IS_END    = "2025-07-31"
OOS_START = "2025-08-01"


# ── 設定 ──────────────────────────────────────────────────────────────────────

@dataclass
class Config:
    signal_tf_min:    int   = 15      # シグナル足（分）
    sma_period:       int   = 20      # SMA期間
    slope_window:     int   = 3       # 傾き計算バー数
    flat_threshold:   float = 0.03    # 平行判定閾値（円）
    min_flat_streak:  int   = 6       # 連続flat必要本数（15分×6=90分）
    swing_lookback:   int   = 20      # スウィング参照バー数（15分×20=300分）
    tp_pips:          float = 0.05    # 固定TP（円）= 5pips
    fixed_stop_pips:  float = 0.02    # 固定ストップ（円）= 2pips (0=スウィング使用)
    event_buffer_min: int   = 60      # イベント回避（分）
    spread:           float = 0.002   # 片道スプレッド（円）= 0.2pips
    risk_pct:         float = 0.5     # 1トレードリスク（資金の%）
    initial_capital:  float = 1_000_000
    cooldown_bars:    int   = 8       # クールダウン（15分×8=120分）


# ── データ ─────────────────────────────────────────────────────────────────────

def load_price(path: Path) -> pd.DataFrame:
    df = pd.read_csv(path, header=None,
                     names=["date","time","open","high","low","close","vol"])
    df["dt"] = pd.to_datetime(df["date"] + " " + df["time"], format="%Y.%m.%d %H:%M")
    df = df.set_index("dt").sort_index()
    for c in ["open","high","low","close"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df.dropna(subset=["close"])


def resample_tf(df1m: pd.DataFrame, tf_min: int) -> pd.DataFrame:
    rule = f"{tf_min}min"
    return df1m.resample(rule, closed="left", label="left").agg(
        open=("open","first"), high=("high","max"),
        low=("low","min"),   close=("close","last"),
    ).dropna(subset=["close"])


def load_events(path: Path) -> list[datetime]:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    pat = re.compile(r"(\d{4}/\d{2}/\d{2})\([^)]+\)\s+(\d{2}:\d{2})")
    events = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        m = pat.match(str(row[7]) if row[7] else "")
        if m:
            try:
                events.append(datetime.strptime(
                    f"{m.group(1)} {m.group(2)}", "%Y/%m/%d %H:%M"))
            except ValueError:
                pass
    print(f"  イベント: {len(events)}件")
    return events


def build_event_mask(index: pd.DatetimeIndex,
                     events: list[datetime], buf_min: int) -> pd.Series:
    mask = pd.Series(False, index=index)
    buf  = timedelta(minutes=buf_min)
    for ev in events:
        ts = pd.Timestamp(ev)
        mask[(index >= ts - buf) & (index <= ts + buf)] = True
    return mask


# ── インジケーター ─────────────────────────────────────────────────────────────

def calc_indicators(df: pd.DataFrame, cfg: Config) -> pd.DataFrame:
    df = df.copy()
    df["sma"]   = df["close"].rolling(cfg.sma_period).mean()
    df["slope"] = df["sma"] - df["sma"].shift(cfg.slope_window)
    df["is_flat"] = df["slope"].abs() < cfg.flat_threshold

    # 連続flat本数
    streak = (df["is_flat"]
              .groupby((~df["is_flat"]).cumsum())
              .cumcount() + 1).where(df["is_flat"], 0)
    df["ready"] = streak >= cfg.min_flat_streak

    # 日足始値（日次フィルター用）
    date_col   = df.index.normalize()
    daily_open = df.groupby(date_col)["open"].first()
    df["daily_open"] = date_col.map(daily_open)

    # SMAクロス
    above = df["close"] > df["sma"]
    df["cross_up"] = (~above.shift(1).fillna(False)) & above
    df["cross_dn"] =  above.shift(1).fillna(False)  & (~above)
    return df


# ── トレード ───────────────────────────────────────────────────────────────────

@dataclass
class Trade:
    direction:   str
    entry_dt:    pd.Timestamp
    entry_price: float
    stop_price:  float
    tp_price:    float
    units:       float
    entry_bar:   int                    = 0
    exit_dt:     Optional[pd.Timestamp] = None
    exit_price:  Optional[float]        = None
    exit_reason: str                    = ""
    pnl_jpy:     float                  = 0.0

    @property
    def closed(self) -> bool:
        return self.exit_dt is not None


# ── バックテスト ───────────────────────────────────────────────────────────────

def run(df: pd.DataFrame, cfg: Config, events: list[datetime]) -> list[Trade]:
    ev_mask    = build_event_mask(df.index, events, cfg.event_buffer_min)
    trades     = []
    position: Optional[Trade] = None
    capital    = cfg.initial_capital

    idx        = df.index
    arr_close  = df["close"].values
    arr_high   = df["high"].values
    arr_low    = df["low"].values
    arr_sma    = df["sma"].values
    arr_flat   = df["is_flat"].values
    arr_ready  = df["ready"].values
    arr_cup    = df["cross_up"].values
    arr_cdn    = df["cross_dn"].values
    arr_ev     = ev_mask.values
    arr_dopen  = df["daily_open"].values

    last_bar   = -999
    start      = cfg.sma_period + cfg.min_flat_streak + cfg.slope_window + 2

    for i in range(start, len(df)):
        bar_dt = idx[i]
        close  = arr_close[i]
        high   = arr_high[i]
        low    = arr_low[i]

        if np.isnan(arr_sma[i]):
            continue

        # ── 決済チェック ──────────────────────────────────────────────────────
        if position is not None:
            reason     = None
            exit_price = close

            if position.direction == "long":
                if high >= position.tp_price:
                    reason, exit_price = "tp",   position.tp_price - cfg.spread
                elif low <= position.stop_price:
                    reason, exit_price = "stop",  position.stop_price - cfg.spread
            else:
                if low <= position.tp_price:
                    reason, exit_price = "tp",   position.tp_price + cfg.spread
                elif high >= position.stop_price:
                    reason, exit_price = "stop",  position.stop_price + cfg.spread

            if reason:
                pnl = position.units * (exit_price - position.entry_price)
                if position.direction == "short":
                    pnl = -pnl
                position.exit_dt     = bar_dt
                position.exit_price  = exit_price
                position.exit_reason = reason
                position.pnl_jpy     = pnl
                capital             += pnl
                last_bar             = i
                trades.append(position)
                position = None

        # ── エントリーチェック ────────────────────────────────────────────────
        if position is not None:
            continue
        if arr_ev[i]:
            continue
        if i - last_bar < cfg.cooldown_bars:
            continue
        if not arr_ready[i - 1]:          # 直前バーまでに準備完了
            continue

        d_open = arr_dopen[i]
        if np.isnan(d_open):
            continue

        sw = min(cfg.swing_lookback, i)
        direction = None

        # 日足始値フィルター + SMAクロス
        if arr_cup[i] and close > d_open:
            direction = "long"
        elif arr_cdn[i] and close < d_open:
            direction = "short"

        if direction is None:
            continue

        if direction == "long":
            entry_price = close + cfg.spread
            if cfg.fixed_stop_pips > 0:
                stop_price = entry_price - cfg.fixed_stop_pips
            else:
                stop_price = float(np.min(arr_low[i - sw: i + 1]))
            tp_price = entry_price + cfg.tp_pips
        else:
            entry_price = close - cfg.spread
            if cfg.fixed_stop_pips > 0:
                stop_price = entry_price + cfg.fixed_stop_pips
            else:
                stop_price = float(np.max(arr_high[i - sw: i + 1]))
            tp_price = entry_price - cfg.tp_pips

        stop_dist = abs(entry_price - stop_price)
        if stop_dist < 0.001:
            continue

        units = (capital * cfg.risk_pct / 100) / stop_dist
        if units < 1:
            continue

        position = Trade(
            direction=direction, entry_dt=bar_dt,
            entry_price=entry_price, stop_price=stop_price,
            tp_price=tp_price, units=units, entry_bar=i,
        )

    # 強制クローズ
    if position is not None:
        ep  = arr_close[-1]
        ep -= cfg.spread if position.direction == "long" else -cfg.spread
        pnl = position.units * (ep - position.entry_price)
        if position.direction == "short":
            pnl = -pnl
        position.exit_dt     = idx[-1]
        position.exit_price  = ep
        position.exit_reason = "end"
        position.pnl_jpy     = pnl
        trades.append(position)

    return [t for t in trades if t.closed]


# ── 統計 ───────────────────────────────────────────────────────────────────────

def compute_stats(trades: list[Trade], cap0: float) -> dict:
    if not trades:
        return dict(total_trades=0, win_rate=0, pf=0, max_dd=0,
                    total_pnl=0, final_capital=cap0)
    wins   = [t for t in trades if t.pnl_jpy > 0]
    losses = [t for t in trades if t.pnl_jpy <= 0]
    gp     = sum(t.pnl_jpy for t in wins)
    gl     = abs(sum(t.pnl_jpy for t in losses))
    pf     = round(gp / gl, 2) if gl > 0 else float("inf")
    eq, pk, dd = cap0, cap0, 0.0
    for t in sorted(trades, key=lambda x: x.exit_dt):
        eq += t.pnl_jpy; pk = max(pk, eq)
        dd  = max(dd, (pk - eq) / pk * 100)
    reasons = {}
    for t in trades:
        reasons[t.exit_reason] = reasons.get(t.exit_reason, 0) + 1
    holds = [(t.exit_dt - t.entry_dt).total_seconds() / 60 for t in trades]
    return dict(
        total_trades=len(trades), win_rate=round(len(wins)/len(trades)*100,1),
        pf=pf, max_dd=round(dd,2), total_pnl=round(sum(t.pnl_jpy for t in trades)),
        final_capital=round(cap0 + sum(t.pnl_jpy for t in trades)),
        exit_reasons=reasons, avg_hold_min=round(sum(holds)/len(holds),1),
    )


def split_is_oos(trades):
    cut = pd.Timestamp(IS_END)
    is_t  = [t for t in trades if t.entry_dt <= cut and t.exit_reason != "end"]
    oos_t = [t for t in trades if t.entry_dt >  cut and t.exit_reason != "end"]
    return is_t, oos_t


def print_stats(s: dict, label: str):
    print(f"\n  [{label}]")
    print(f"  件数:{s['total_trades']}  勝率:{s['win_rate']}%  "
          f"PF:{s['pf']}  DD:{s['max_dd']}%  "
          f"損益:{s['total_pnl']:+,.0f}円  平均保有:{s.get('avg_hold_min',0)}分")
    print(f"  決済理由:{s.get('exit_reasons',{})}")


# ── メイン ────────────────────────────────────────────────────────────────────

def main():
    cfg      = Config()
    run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*60}")
    print(f"  USD/JPY バックテスト v3 — 20SMA接線ブレイクアウト")
    print(f"  実行:{run_time}  足:{cfg.signal_tf_min}分  "
          f"flat閾値:{cfg.flat_threshold}円/{cfg.slope_window}本  "
          f"連続:{cfg.min_flat_streak}本")
    print(f"  TP:{cfg.tp_pips}円({cfg.tp_pips*100:.0f}pips)  "
          f"swing:{cfg.swing_lookback}本  "
          f"スプレッド:{cfg.spread}円  risk:{cfg.risk_pct}%")
    print(f"  日足フィルター: 始値より上→ロング / 下→ショート")
    print(f"{'='*60}")

    print("\nデータ読み込み中...")
    df1m = load_price(PRICE_CSV)
    print(f"  1分足:{len(df1m):,}本  {df1m.index[0]} ~ {df1m.index[-1]}")
    df   = resample_tf(df1m, cfg.signal_tf_min)
    print(f"  {cfg.signal_tf_min}分足:{len(df):,}本")

    events = load_events(EVENTS_XLS)

    print("\nインジケーター計算中...")
    df = calc_indicators(df, cfg)

    print("\nバックテスト実行中...")
    all_trades = run(df, cfg, events)
    print(f"  総トレード:{len(all_trades)}件")

    is_t, oos_t = split_is_oos(all_trades)
    is_s  = compute_stats(is_t,  cfg.initial_capital)
    oos_s = compute_stats(oos_t, cfg.initial_capital)

    print(f"\n{'='*60}")
    print(f"  IS  [~{IS_END}]")
    print_stats(is_s, "IS")
    print(f"\n  OOS [{OOS_START}~]")
    print_stats(oos_s, "OOS")
    print(f"{'='*60}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _save(trades, label):
        if not trades:
            return
        path = OUTPUT_DIR / f"usdjpy_{label}_{ts}.csv"
        pd.DataFrame([{
            "entry_dt": str(t.entry_dt)[:16], "direction": t.direction,
            "entry_price": t.entry_price, "stop_price": t.stop_price,
            "tp_price": t.tp_price, "units": round(t.units),
            "exit_dt": str(t.exit_dt)[:16], "exit_price": t.exit_price,
            "exit_reason": t.exit_reason, "pnl_jpy": round(t.pnl_jpy),
        } for t in trades]).to_csv(path, index=False, encoding="utf-8-sig")
        print(f"  {label.upper()} CSV: {path.name}")

    _save(is_t,  "is")
    _save(oos_t, "oos")
    print(f"\n{'='*60}\n")

    return is_s, oos_s, all_trades


if __name__ == "__main__":
    main()
