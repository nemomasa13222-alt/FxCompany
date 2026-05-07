# -*- coding: utf-8 -*-
"""
USD/JPY 1分足バックテスト — 20SMA平均傾きゼロクロス 転換戦略
生傾き=SMA.diff(3)  平均傾き=生傾きの5本MA  ゼロクロスでエントリー
TP:10pips / SL:5pips / スプレッド:0.2pips
実行: python backtest/usdjpy_slope_reversal.py
"""
from __future__ import annotations
import sys, re
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional
import pandas as pd
import numpy as np

BASE_DIR   = Path(__file__).parent
PRICE_CSV  = BASE_DIR / "USDJPY" / "USDJPY__20241014~20260501（1分足）.csv"
EVENTS_XLS = BASE_DIR.parent / "米国経済指標カレンダー_2023Dec-2026May.xlsx"
OUTPUT_DIR = BASE_DIR / "results"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
IS_END = "2025-07-31"
OOS_START = "2025-08-01"

@dataclass
class Config:
    signal_tf_min:     int   = 15
    sma_period:        int   = 20
    slope_window:      int   = 3
    smooth_period:     int   = 5
    min_slope_confirm: float = 0.05   # ゼロクロス後この傾きを超えたらエントリー
    max_pending_bars:  int   = 5      # 確認を待つ最大バー数（超えたらキャンセル）
    tp_pips:           float = 0.10
    sl_pips:           float = 0.05
    spread:            float = 0.002
    risk_pct:          float = 0.5
    initial_capital:   float = 1_000_000
    event_buffer_min:  int   = 60
    cooldown_bars:     int   = 8

@dataclass
class Trade:
    direction:   str
    entry_dt:    pd.Timestamp
    entry_price: float
    stop_price:  float
    tp_price:    float
    units:       float
    entry_bar:   int = 0
    exit_dt:     Optional[pd.Timestamp] = None
    exit_price:  Optional[float] = None
    exit_reason: str = ""
    pnl_jpy:     float = 0.0
    @property
    def closed(self): return self.exit_dt is not None

def load_price(path):
    df = pd.read_csv(path, header=None,
                     names=["date","time","open","high","low","close","vol"])
    df["dt"] = pd.to_datetime(df["date"]+" "+df["time"], format="%Y.%m.%d %H:%M")
    df = df.set_index("dt").sort_index()
    for c in ["open","high","low","close"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df.dropna(subset=["close"])

def resample_tf(df1m, tf_min):
    return df1m.resample(f"{tf_min}min", closed="left", label="left").agg(
        open=("open","first"), high=("high","max"),
        low=("low","min"), close=("close","last")).dropna(subset=["close"])

def load_events(path):
    from openpyxl import load_workbook
    wb = load_workbook(path); ws = wb.active
    pat = re.compile(r"(\d{4}/\d{2}/\d{2})\([^)]+\)\s+(\d{2}:\d{2})")
    evts = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        m = pat.match(str(row[7]) if row[7] else "")
        if m:
            try: evts.append(datetime.strptime(f"{m.group(1)} {m.group(2)}", "%Y/%m/%d %H:%M"))
            except: pass
    print(f"  イベント: {len(evts)}件"); return evts

def build_event_mask(index, events, buf_min):
    mask = pd.Series(False, index=index); buf = timedelta(minutes=buf_min)
    for ev in events:
        ts = pd.Timestamp(ev)
        mask[(index >= ts-buf) & (index <= ts+buf)] = True
    return mask

def calc_indicators(df, cfg):
    df = df.copy()
    df["sma"]          = df["close"].rolling(cfg.sma_period).mean()
    df["raw_slope"]    = df["sma"].diff(cfg.slope_window)
    df["smooth_slope"] = df["raw_slope"].rolling(cfg.smooth_period).mean()
    ss = df["smooth_slope"]; lag = ss.shift(1)
    df["cross_up"] = (lag < 0) & (ss >= 0)
    df["cross_dn"] = (lag > 0) & (ss <= 0)
    return df

def run(df, cfg, events):
    ev_mask   = build_event_mask(df.index, events, cfg.event_buffer_min)
    trades    = []; position = None; capital = cfg.initial_capital
    idx       = df.index
    arr_close = df["close"].values; arr_high = df["high"].values
    arr_low   = df["low"].values;   arr_sma  = df["sma"].values
    arr_cup   = df["cross_up"].values; arr_cdn = df["cross_dn"].values
    arr_ss    = df["smooth_slope"].values
    arr_ev    = ev_mask.values; last_bar = -999
    start     = cfg.sma_period + cfg.slope_window + cfg.smooth_period + 2

    # 確認待ち状態
    pending_dir = None   # "long" or "short" or None
    pending_bar = -999   # ゼロクロスが起きたバー

    for i in range(start, len(df)):
        bar_dt = idx[i]; close = arr_close[i]
        high = arr_high[i]; low = arr_low[i]
        ss   = arr_ss[i]
        if np.isnan(arr_sma[i]) or np.isnan(ss): continue

        # ── 決済チェック ──────────────────────────────────────────────────────
        if position is not None:
            reason = None; exit_price = close
            if position.direction == "long":
                if high >= position.tp_price:
                    reason, exit_price = "tp",   position.tp_price - cfg.spread
                elif low <= position.stop_price:
                    reason, exit_price = "stop", position.stop_price - cfg.spread
            else:
                if low <= position.tp_price:
                    reason, exit_price = "tp",   position.tp_price + cfg.spread
                elif high >= position.stop_price:
                    reason, exit_price = "stop", position.stop_price + cfg.spread
            if reason:
                pnl = position.units * (exit_price - position.entry_price)
                if position.direction == "short": pnl = -pnl
                position.exit_dt = bar_dt; position.exit_price = exit_price
                position.exit_reason = reason; position.pnl_jpy = pnl
                capital += pnl; last_bar = i
                trades.append(position); position = None
                pending_dir = None  # ポジション決済でpendingもリセット

        if position is not None: continue
        if arr_ev[i]: continue
        if i - last_bar < cfg.cooldown_bars: continue

        # ── 新ゼロクロス検出 → pending開始 ──────────────────────────────────
        if arr_cup[i]:
            pending_dir = "long";  pending_bar = i
        elif arr_cdn[i]:
            pending_dir = "short"; pending_bar = i

        # ── pending中: 確認チェック ───────────────────────────────────────────
        if pending_dir is not None:
            bars_waiting = i - pending_bar

            # タイムアウト or 傾きが逆に戻ったらキャンセル
            if bars_waiting > cfg.max_pending_bars:
                pending_dir = None; continue
            if pending_dir == "long"  and ss < 0:
                pending_dir = None; continue  # 傾きが再びマイナスへ→偽シグナル
            if pending_dir == "short" and ss > 0:
                pending_dir = None; continue  # 傾きが再びプラスへ→偽シグナル

            # 傾きが min_slope_confirm を超えたらエントリー確定
            confirmed = False
            if pending_dir == "long"  and ss >= cfg.min_slope_confirm:
                confirmed = True
            if pending_dir == "short" and ss <= -cfg.min_slope_confirm:
                confirmed = True

            if not confirmed:
                continue

            direction  = pending_dir
            pending_dir = None  # リセット

            if direction == "long":
                ep = close + cfg.spread; sp = ep - cfg.sl_pips; tp = ep + cfg.tp_pips
            else:
                ep = close - cfg.spread; sp = ep + cfg.sl_pips; tp = ep - cfg.tp_pips

            units = (capital * cfg.risk_pct / 100) / cfg.sl_pips
            if units < 1: continue
            position = Trade(direction=direction, entry_dt=bar_dt,
                             entry_price=ep, stop_price=sp, tp_price=tp,
                             units=units, entry_bar=i)

    if position is not None:
        ep = arr_close[-1]
        ep -= cfg.spread if position.direction == "long" else -cfg.spread
        pnl = position.units * (ep - position.entry_price)
        if position.direction == "short": pnl = -pnl
        position.exit_dt = idx[-1]; position.exit_price = ep
        position.exit_reason = "end"; position.pnl_jpy = pnl
        trades.append(position)

    return [t for t in trades if t.closed]

def compute_stats(trades, cap0):
    if not trades:
        return dict(total_trades=0,win_rate=0,pf=0,max_dd=0,
                    total_pnl=0,final_capital=cap0,long_trades=0,short_trades=0)
    wins = [t for t in trades if t.pnl_jpy > 0]
    loss = [t for t in trades if t.pnl_jpy <= 0]
    gp = sum(t.pnl_jpy for t in wins); gl = abs(sum(t.pnl_jpy for t in loss))
    pf = round(gp/gl, 2) if gl > 0 else float("inf")
    eq, pk, dd = cap0, cap0, 0.0
    for t in sorted(trades, key=lambda x: x.exit_dt):
        eq += t.pnl_jpy; pk = max(pk, eq)
        dd = max(dd, (pk-eq)/pk*100)
    reasons = {}
    for t in trades: reasons[t.exit_reason] = reasons.get(t.exit_reason, 0) + 1
    holds = [(t.exit_dt-t.entry_dt).total_seconds()/60 for t in trades]
    return dict(
        total_trades=len(trades), win_rate=round(len(wins)/len(trades)*100,1),
        pf=pf, max_dd=round(dd,2), total_pnl=round(sum(t.pnl_jpy for t in trades)),
        final_capital=round(cap0+sum(t.pnl_jpy for t in trades)),
        exit_reasons=reasons, avg_hold_min=round(sum(holds)/len(holds),1),
        long_trades=sum(1 for t in trades if t.direction=="long"),
        short_trades=sum(1 for t in trades if t.direction=="short"),
    )

def split_is_oos(trades):
    cut = pd.Timestamp(IS_END)
    return ([t for t in trades if t.entry_dt<=cut and t.exit_reason!="end"],
            [t for t in trades if t.entry_dt>cut  and t.exit_reason!="end"])

def print_stats(s, label):
    print(f"\n  [{label}]")
    print(f"  件数:{s['total_trades']}  Long:{s.get('long_trades',0)} Short:{s.get('short_trades',0)}  "
          f"勝率:{s['win_rate']}%  PF:{s['pf']}  DD:{s['max_dd']}%")
    print(f"  損益:{s['total_pnl']:+,.0f}円  最終資金:{s['final_capital']:,.0f}円  "
          f"平均保有:{s.get('avg_hold_min',0)}分")
    print(f"  決済:{s.get('exit_reasons',{})}")

def main():
    cfg = Config()
    print(f"\n{'='*60}")
    print(f"  USD/JPY — 20SMA平均傾きゼロクロス転換戦略")
    print(f"  足:{cfg.signal_tf_min}分  SMA:{cfg.sma_period}  "
          f"傾きWindow:{cfg.slope_window}本  平均化:{cfg.smooth_period}本")
    print(f"  TP:{cfg.tp_pips*100:.0f}pips  SL:{cfg.sl_pips*100:.0f}pips  "
          f"スプレッド:{cfg.spread*100:.1f}pips  risk:{cfg.risk_pct}%")
    print(f"  確認閾値:{cfg.min_slope_confirm}  最大待機:{cfg.max_pending_bars}本")
    print(f"  損益分岐勝率: {cfg.sl_pips/(cfg.sl_pips+cfg.tp_pips)*100:.1f}%")
    print(f"{'='*60}")

    print("\nデータ読み込み中...")
    df1m = load_price(PRICE_CSV)
    df   = resample_tf(df1m, cfg.signal_tf_min)
    print(f"  1分足:{len(df1m):,}本  →  {cfg.signal_tf_min}分足:{len(df):,}本")
    print(f"  期間: {df.index[0]} ~ {df.index[-1]}")

    events = load_events(EVENTS_XLS)

    print("\nインジケーター計算中...")
    df = calc_indicators(df, cfg)
    n_up = int(df["cross_up"].sum()); n_dn = int(df["cross_dn"].sum())
    print(f"  シグナル: ロング{n_up}件 / ショート{n_dn}件 = 計{n_up+n_dn}件")

    print("\nバックテスト実行中...")
    all_trades = run(df, cfg, events)
    print(f"  総トレード:{len(all_trades)}件")

    is_t, oos_t = split_is_oos(all_trades)
    is_s  = compute_stats(is_t,  cfg.initial_capital)
    oos_s = compute_stats(oos_t, cfg.initial_capital)

    print(f"\n{'='*60}")
    print_stats(is_s,  f"IS  [~{IS_END}]")
    print_stats(oos_s, f"OOS [{OOS_START}~]")
    print(f"{'='*60}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    def _save(trades, label):
        if not trades: return
        path = OUTPUT_DIR / f"usdjpy_slope_{label}_{ts}.csv"
        pd.DataFrame([{
            "entry_dt":str(t.entry_dt)[:16], "direction":t.direction,
            "entry_price":round(t.entry_price,3), "stop_price":round(t.stop_price,3),
            "tp_price":round(t.tp_price,3), "units":round(t.units),
            "exit_dt":str(t.exit_dt)[:16], "exit_price":round(t.exit_price,3),
            "exit_reason":t.exit_reason, "pnl_jpy":round(t.pnl_jpy),
        } for t in trades]).to_csv(path, index=False, encoding="utf-8-sig")
        print(f"  {label.upper()} CSV: {path.name}")
    _save(is_t, "is"); _save(oos_t, "oos")
    print(f"\n{'='*60}\n")
    return is_s, oos_s

if __name__ == "__main__":
    main()
